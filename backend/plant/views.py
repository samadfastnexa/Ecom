from decimal import Decimal
from datetime import timedelta

from django.contrib.auth.models import User
from django.db.models import Sum, Count
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.permissions import SAFE_METHODS
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import DeliveryRecord, CustomerType, BottleType, PlantSettings
from .serializers import (
    DeliveryRecordSerializer, CustomerTypeSerializer, BottleTypeSerializer,
)
from .permissions import PlantModelPermission, CanAccessPlant, IsPlantAdmin


def _apply_filters(qs, params):
    """Filter by date range + payment_status + customer_type + bottle_type."""
    date = params.get('date')
    start = params.get('start')
    end = params.get('end')
    if date:
        qs = qs.filter(date=date)
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)

    payment_status = params.get('payment_status')
    if payment_status == 'paid':
        qs = qs.filter(paid=True)
    elif payment_status == 'partial':
        qs = qs.filter(paid=False, paid_amount__gt=0)
    elif payment_status == 'unpaid':
        qs = qs.filter(paid_amount=0)

    customer_type = params.get('customer_type')
    if customer_type:
        qs = qs.filter(customer_type_id=customer_type)

    bottle_type = params.get('bottle_type')
    if bottle_type:
        qs = qs.filter(bottle_type_id=bottle_type)

    return qs


# Keep old name as alias so nothing else breaks.
_apply_date_filters = _apply_filters


class DeliveryRecordViewSet(viewsets.ModelViewSet):
    serializer_class = DeliveryRecordSerializer
    permission_classes = [PlantModelPermission]
    queryset = DeliveryRecord.objects.select_related('customer').all()

    def get_queryset(self):
        return _apply_filters(super().get_queryset(), self.request.query_params)


class _TypeViewSetMixin:
    """Shared behaviour for the customer-type and bottle-type endpoints:
    reads for any plant user, writes for admins, optional ?active=true filter."""

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.query_params.get('active') == 'true':
            qs = qs.filter(is_active=True)
        return qs

    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [CanAccessPlant()]
        return [IsPlantAdmin()]


class CustomerTypeViewSet(_TypeViewSetMixin, viewsets.ModelViewSet):
    serializer_class = CustomerTypeSerializer
    queryset = CustomerType.objects.all()


class BottleTypeViewSet(_TypeViewSetMixin, viewsets.ModelViewSet):
    serializer_class = BottleTypeSerializer
    queryset = BottleType.objects.all()


@api_view(['GET'])
@permission_classes([CanAccessPlant])
def plant_summary(request):
    """Totals for the selected range; defaults to today's collection."""
    params = request.query_params
    qs = _apply_filters(DeliveryRecord.objects.all(), params)
    # default to today when no date filter at all
    if not any(params.get(k) for k in ('date', 'start', 'end', 'payment_status', 'customer_type', 'bottle_type')):
        qs = qs.filter(date=timezone.localdate())

    agg = qs.aggregate(
        records=Count('id'),
        bottles=Coalesce(Sum('bottles'), 0),
        amount=Coalesce(Sum('amount'), Decimal('0')),
        paid_amount=Coalesce(Sum('paid_amount'), Decimal('0')),
    )
    pending = agg['amount'] - agg['paid_amount']
    agg['pending'] = pending if pending > 0 else Decimal('0')
    agg['unpaid_amount'] = agg['pending']  # backward-compatible alias
    agg['houses'] = qs.values('house').distinct().count()
    return Response(agg)


@api_view(['GET'])
@permission_classes([CanAccessPlant])
def plant_customers(request):
    """Registered customers available for linking, with their effective price."""
    standard = PlantSettings.load().standard_unit_price
    users = (
        User.objects.filter(profile__user_type='customer')
        .select_related('profile')
        .order_by('username')
    )
    data = []
    for u in users:
        profile = getattr(u, 'profile', None)
        custom = profile.custom_bottle_price if profile else None
        data.append({
            'id': u.id,
            'username': u.username,
            'name': u.get_full_name() or u.username,
            'address': getattr(profile, 'address', None),
            'phone': getattr(profile, 'phone_number', None),
            'price': custom if custom is not None else standard,
            'has_custom_price': custom is not None,
        })
    return Response(data)


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([CanAccessPlant])
def plant_settings(request):
    """Read or update the standard per-bottle price."""
    obj = PlantSettings.load()
    if request.method in ('PUT', 'PATCH'):
        user = request.user
        can_edit = (
            user.is_superuser
            or user.is_staff
            or user.has_perm('plant.change_plantsettings')
        )
        if not can_edit:
            return Response(
                {'detail': 'You do not have permission to change settings.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        price = request.data.get('standard_unit_price')
        if price is not None:
            obj.standard_unit_price = price
            obj.save()
    return Response({'standard_unit_price': obj.standard_unit_price})


@api_view(['GET'])
@permission_classes([CanAccessPlant])
def plant_analytics(request):
    """Aggregated analytics for the admin dashboard (defaults to last 30 days)."""
    params = request.query_params
    qs = _apply_filters(DeliveryRecord.objects.all(), params)
    if not any(params.get(k) for k in ('date', 'start', 'end', 'payment_status', 'customer_type', 'bottle_type')):
        qs = qs.filter(date__gte=timezone.localdate() - timedelta(days=29))

    totals = qs.aggregate(
        records=Count('id'),
        bottles=Coalesce(Sum('bottles'), 0),
        amount=Coalesce(Sum('amount'), Decimal('0')),
        paid_amount=Coalesce(Sum('paid_amount'), Decimal('0')),
    )
    pending = totals['amount'] - totals['paid_amount']
    totals['pending'] = pending if pending > 0 else Decimal('0')
    totals['unpaid_amount'] = totals['pending']

    daily = list(
        qs.values('date')
        .annotate(
            bottles=Sum('bottles'),
            amount=Sum('amount'),
            paid_amount=Sum('paid_amount'),
            records=Count('id'),
        )
        .order_by('date')
    )
    top_houses = list(
        qs.values('house')
        .annotate(bottles=Sum('bottles'), amount=Sum('amount'))
        .order_by('-amount')[:8]
    )
    return Response({'totals': totals, 'daily': daily, 'top_houses': top_houses})


@api_view(['GET'])
@permission_classes([CanAccessPlant])
def plant_export(request):
    """Export the (optionally date-filtered) records as an .xlsx workbook."""
    params = request.query_params
    qs = _apply_filters(
        DeliveryRecord.objects.select_related(
            'customer', 'customer_type', 'bottle_type'
        ),
        params,
    ).order_by('date', 'id')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deliveries'

    headers = [
        'Date', 'House / Address', 'Customer Account', 'Customer Type',
        'Bottle Type', 'Bottles', 'Unit Price', 'Amount',
        'Received', 'Pending', 'Status', 'Notes',
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='088395')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    total_bottles = 0
    total_amount = Decimal('0')
    total_received = Decimal('0')
    total_pending = Decimal('0')
    for r in qs:
        customer = (
            (r.customer.get_full_name() or r.customer.username)
            if r.customer else ''
        )
        ws.append([
            r.date.strftime('%Y-%m-%d'),
            r.house,
            customer,
            r.customer_type.name if r.customer_type else '',
            r.bottle_type.name if r.bottle_type else '',
            r.bottles,
            float(r.unit_price),
            float(r.amount),
            float(r.paid_amount),
            float(r.pending),
            r.payment_status.title(),
            r.notes,
        ])
        total_bottles += r.bottles
        total_amount += r.amount
        total_received += r.paid_amount
        total_pending += r.pending

    # Totals row
    ws.append([])
    total_row = [
        '', '', '', '', 'TOTAL', total_bottles, '',
        float(total_amount), float(total_received), float(total_pending), '', '',
    ]
    ws.append(total_row)
    last = ws.max_row
    for col in (5, 6, 8, 9, 10):
        ws.cell(row=last, column=col).font = Font(bold=True)

    # Reasonable column widths
    widths = [12, 26, 20, 16, 14, 9, 11, 12, 12, 12, 10, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Filename reflects the date filter
    start = params.get('start')
    end = params.get('end')
    date = params.get('date')
    if date:
        suffix = date
    elif start or end:
        suffix = f"{start or 'start'}_to_{end or 'end'}"
    else:
        suffix = timezone.localdate().strftime('%Y-%m-%d')
    filename = f"plant-deliveries-{suffix}.xlsx"

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response
